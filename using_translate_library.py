import openpyxl
from translate import Translator
import json
import os

# Configure the translator object
translator = Translator(from_lang="mr", to_lang="en")

# Define the path to your Excel file and checkpoint file
INPUT_FILE = 'null.xlsx'
OUTPUT_FILE = 'using_translate_library_translated_output_excel_file.xlsx'
CHECKPOINT_FILE = 'translation_checkpoint.json'

# Function to translate text using the `translate` library
def translate_text(text):
    try:
        return translator.translate(text)
    except Exception as e:
        print(f"An error occurred during translation: {e}")
        return text  # Return the original text if translation fails

# Load or initialize the checkpoint
if os.path.exists(CHECKPOINT_FILE):
    with open(CHECKPOINT_FILE, 'r') as f:
        checkpoint = json.load(f)
else:
    checkpoint = {}

# Load the workbook
wb = openpyxl.load_workbook(INPUT_FILE)

# Go through each sheet in the workbook
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    start_row = checkpoint.get(sheet_name, 1)
    
    # Go through each cell in the column
    for row in ws.iter_rows(min_row=start_row, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                # Translate and update the cell value
                cell.value = translate_text(cell.value)
                print(f"Translated cell {cell.coordinate} in sheet '{sheet_name}' to: {cell.value}")
    
    # Update the checkpoint after each sheet is processed
    checkpoint[sheet_name] = ws.max_row + 1
    with open(CHECKPOINT_FILE, 'w') as f:
        json.dump(checkpoint, f)

    print(f"Completed processing sheet: {sheet_name}")

# Save the translated workbook
wb.save(OUTPUT_FILE)
print("Translation completed and workbook saved.")
