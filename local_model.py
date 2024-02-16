from transformers import MarianMTModel, MarianTokenizer
import openpyxl

# Specify the model name
model_name = 'Helsinki-NLP/opus-mt-mr-en'  # Replace with the correct model

# Load the tokenizer and model
tokenizer = MarianTokenizer.from_pretrained(model_name)
model = MarianMTModel.from_pretrained(model_name)

# Function to translate text
def translate(text):
    # Tokenize the text
    batch = tokenizer.prepare_seq2seq_batch(src_texts=[text])
    # Perform translation and decode the output
    translated = model.generate(**batch)
    return tokenizer.decode(translated[0], skip_special_tokens=True)

# Load the Excel workbook
wb = openpyxl.load_workbook('null.xlsx')
ws = wb.active

# Translate each cell
for row in ws.iter_rows(min_row=1, max_col=1, max_row=ws.max_row):
    for cell in row:
        if cell.value:
            translated_text = translate(cell.value)
            print(f'Translated: {translated_text}')
            cell.value = translated_text

# Save the translated workbook
wb.save('model_translated_excel_file.xlsx')
print('Translation completed and workbook saved.')
