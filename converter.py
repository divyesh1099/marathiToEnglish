import openpyxl
import requests
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
import json
import os
import time
from customDivyeshAzureKeys import SubscriptionKey

# Constants for API and processing
MAX_WORKERS = 2  # Adjust based on your API rate limits and your machine capability
BATCH_SIZE = 2  # Adjust based on the maximum batch size your API plan allows

# Azure credentials and endpoint
SUBSCRIPTION_KEY = SubscriptionKey
ENDPOINT = "https://api.cognitive.microsofttranslator.com"
LOCATION = 'centralus'

# Paths to your files
INPUT_FILE = 'null.xlsx'
OUTPUT_FILE = 'translated_excel_file.xlsx'
CHECKPOINT_FILE = 'checkpoint.json'

# Function to translate a batch of texts
def translate_batch(texts, subscription_key, endpoint, location, from_lang='mr', to_lang='en', retries=3, backoff_factor=0.5):
    path = '/translate'
    constructed_url = endpoint + path

    headers = {
        'Ocp-Apim-Subscription-Key': subscription_key,
        'Ocp-Apim-Subscription-Region': location,
        'Content-type': 'application/json',
        'X-ClientTraceId': str(uuid.uuid4())
    }

    body = [{'text': text} for text in texts]

    params = {
        'api-version': '3.0',
        'from': from_lang,
        'to': to_lang
    }

    for attempt in range(retries):
        response = requests.post(constructed_url, params=params, headers=headers, json=body)
        if response.status_code == 429:
            if attempt < retries - 1:
                wait_time = (2 ** attempt) * backoff_factor
                print(f"Rate limit hit. Waiting for {wait_time} seconds.")
                time.sleep(wait_time)
            else:
                print("Max retries hit. Raising exception.")
                response.raise_for_status()
        else:
            response.raise_for_status()
            return [item['translations'][0]['text'] for item in response.json()]

# Function to process batches in parallel and update the sheet
def process_sheet(sheet_name, subscription_key, endpoint, location, wb, checkpoint):
    ws = wb[sheet_name]
    row_num = checkpoint.get(sheet_name, 1)
    batch_texts = []
    translations = []

    for row in ws.iter_rows(min_row=row_num, max_col=ws.max_column, max_row=ws.max_row, values_only=True):
        for cell in row:
            if isinstance(cell, str):
                batch_texts.append(cell)
                if len(batch_texts) >= BATCH_SIZE:
                    translations.extend(translate_batch(batch_texts, subscription_key, endpoint, location))
                    batch_texts = []
        if batch_texts:
            translations.extend(translate_batch(batch_texts, subscription_key, endpoint, location))

    update_rows = ws.iter_rows(min_row=row_num, max_col=1, max_row=row_num+len(translations)-1)
    for row, translation in zip(update_rows, translations):
        row[0].value = translation

    checkpoint[sheet_name] = ws.max_row + 1
    print(f"Sheet {sheet_name} processed up to row {checkpoint[sheet_name]}.")

# Load or initialize the checkpoint
if os.path.exists(CHECKPOINT_FILE):
    with open(CHECKPOINT_FILE, 'r') as f:
        checkpoint = json.load(f)
else:
    checkpoint = {}

# Load the workbook
wb = openpyxl.load_workbook(INPUT_FILE)

# Set up ThreadPoolExecutor for parallel processing
with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
    futures = {executor.submit(process_sheet, sheet_name, SUBSCRIPTION_KEY, ENDPOINT, LOCATION, wb, checkpoint): sheet_name for sheet_name in wb.sheetnames}
    for future in as_completed(futures):
        sheet_name = futures[future]
        try:
            future.result()
            print(f"Sheet '{sheet_name}' processing completed.")
        except Exception as exc:
            print(f"An error occurred while processing sheet '{sheet_name}': {exc}")
            print("Any unprocessed or partially processed sheets will need to be retried.")

#Save the final workbook and checkpoint
wb.save(OUTPUT_FILE)
with open(CHECKPOINT_FILE, 'w') as f:
    json.dump(checkpoint, f)

print("Translation completed and workbook saved.")